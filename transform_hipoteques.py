"""
transform_hipoteques.py
Taules INE 13896 (constituides + capital) i 13902 (cancelades)
→ Excel format Idescat: Full de càrrega + CAT + ESP
"""

import urllib.request
import json
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import re

TAULES = {"constituides": "13896", "cancelades": "13902"}
N_PERIODES = 12
GEO_FILTRES = {"CAT": "Cataluña", "ESP": "Total Nacional"}
OUTPUT_FILE = "hipoteques_idescat.xlsx"

# Cada entrada: (paraules clau que han d'aparèixer al nom raw, nom visible)
# Les paraules clau s'usen per cerca parcial case-insensitive
MAPEIG_CONST = [
    (["Total fincas", "Número"],       "Hipoteques constituïdes (nombre)"),
    (["fincas urbanas", "Número"],     "Finques urbanes"),
    (["Viviendas", "Número"],          "Habitatges"),
    (["Solares", "Número"],            "Solars"),
    (["Otros", "Número"],              "Altres"),
    (["rústicas", "Número"],           "Finques rústiques"),
    (["Total fincas", "Importe"],      "Capital prestat (milions d'euros)"),
    (["fincas urbanas", "Importe"],    "Finques urbanes (capital)"),
    (["Viviendas", "Importe"],         "Habitatges (capital)"),
    (["Solares", "Importe"],           "Solars (capital)"),
    (["Otros", "Importe"],             "Altres (capital)"),
    (["rústicas", "Importe"],          "Finques rústiques (capital)"),
]

MAPEIG_CANC = [
    (["Total fincas", "cancelada"],          "Hipoteques cancel·lades (nombre)"),
    (["fincas urbanas", "cancelada"],        "Finques urbanes (cancel·lades)"),
    (["Viviendas", "cancelada"],             "Habitatges (cancel·lades)"),
    (["Solares", "cancelada"],               "Solars (cancel·lades)"),
    (["urbanas: otros", "cancelada"],        "Altres (cancel·lades)"),
    (["rústicas", "cancelada"],              "Finques rústiques (cancel·lades)"),
]

CAPS_VISIBLES = [
    "Hipoteques constituïdes (nombre)", "Finques urbanes", "Habitatges", "Solars", "Altres", "Finques rústiques",
    "Capital prestat (milions d'euros)", "Finques urbanes", "Habitatges", "Solars", "Altres", "Finques rústiques",
    "Hipoteques cancel·lades (nombre)", "Finques urbanes", "Habitatges", "Solars", "Altres", "Finques rústiques",
]

def fetch_taula(id_taula, n=12):
    url = f"https://servicios.ine.es/wstempus/js/ES/DATOS_TABLA/{id_taula}?nult={n}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return json.loads(urllib.request.urlopen(req, timeout=30).read())

def timestamp_a_periode(fecha_ms):
    dt = datetime.datetime.utcfromtimestamp(fecha_ms / 1000)
    return f"{dt.year}{str(dt.month).zfill(2)}"

def parse_taula(dades, geo_text):
    """Retorna dict { nom_raw_normalitzat → { periode → valor } }"""
    result = {}
    for serie in dades:
        nom_full = serie.get("Nombre", "")
        if geo_text not in nom_full:
            continue
        # Normalitzar espais múltiples
        nom_raw = re.sub(r'\s+', ' ', nom_full).strip()
        for d in serie.get("Data", []):
            if d.get("Valor") is None:
                continue
            periode = timestamp_a_periode(d["Fecha"])
            if nom_raw not in result:
                result[nom_raw] = {}
            result[nom_raw][periode] = d["Valor"]
    return result

def cerca_clau(raw_dict, paraules_clau):
    """Troba la primera clau del dict que conté totes les paraules clau."""
    for clau in raw_dict:
        clau_lower = clau.lower()
        if all(p.lower() in clau_lower for p in paraules_clau):
            return clau
    return None

def construir_df(dades_const, dades_canc, geo_text):
    raw_const = parse_taula(dades_const, geo_text)
    raw_canc  = parse_taula(dades_canc,  geo_text)

    print(f"  Noms raw constituides+capital:")
    for k in raw_const:
        print(f"    {k}")
    print(f"  Noms raw cancel·lades:")
    for k in raw_canc:
        print(f"    {k}")

    # Obtenir tots els períodes
    periodes = set()
    for d in list(raw_const.values()) + list(raw_canc.values()):
        periodes.update(d.keys())
    periodes = sorted(periodes, reverse=True)[:N_PERIODES]

    data = {p: {} for p in periodes}

    for paraules, nom_visible in MAPEIG_CONST:
        clau = cerca_clau(raw_const, paraules)
        print(f"  CONST '{nom_visible}' → '{clau}'")
        for p in periodes:
            data[p][nom_visible] = raw_const.get(clau, {}).get(p) if clau else None

    for paraules, nom_visible in MAPEIG_CANC:
        clau = cerca_clau(raw_canc, paraules)
        print(f"  CANC '{nom_visible}' → '{clau}'")
        for p in periodes:
            data[p][nom_visible] = raw_canc.get(clau, {}).get(p) if clau else None

    col_order = [m[1] for m in MAPEIG_CONST] + [m[1] for m in MAPEIG_CANC]
    df = pd.DataFrame(data).T
    df.index.name = "periode"
    df = df[col_order]
    df = df.sort_index(ascending=False)
    return df

def afegir_columnes_buides(df, n_buides=3):
    cols = []
    for i, col in enumerate(df.columns):
        cols.append(df[col])
        if i < len(df.columns) - 1:
            for j in range(n_buides):
                cols.append(pd.Series([None]*len(df), index=df.index, name=f"__b{i}_{j}"))
    return pd.concat(cols, axis=1)

def escriure_full_carrega(ws):
    ws.title = "Full de càrrega"
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(color="FFFFFF", bold=True, size=11)
    ws["A1"] = "Full de càrrega"
    ws["A1"].fill = fill
    ws["A1"].font = font
    ws.column_dimensions["A"].width = 30

def escriure_pestanya_dades(ws, df):
    fill_cap = PatternFill("solid", fgColor="1F3864")
    font_cap = Font(color="FFFFFF", bold=True, size=9)

    df_final = afegir_columnes_buides(df, n_buides=3)

    caps = ["Periode"] + [
        "" if str(c).startswith("__b") else str(c)
        for c in df_final.columns
    ]
    ws.append(caps)

    for cell in ws[1]:
        cell.fill = fill_cap
        cell.font = font_cap
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 50

    for periode, row in df_final.iterrows():
        ws.append([periode] + list(row))

    ws.column_dimensions["A"].width = 10
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 12

def main():
    print("Descarregant dades del INE...")
    try:
        dades_const = fetch_taula(TAULES["constituides"], N_PERIODES)
        dades_canc  = fetch_taula(TAULES["cancelades"],   N_PERIODES)
    except Exception as e:
        print(f"ERROR descàrrega: {e}")
        sys.exit(1)

    wb = Workbook()
    escriure_full_carrega(wb.active)

    for nom_geo, geo_text in GEO_FILTRES.items():
        print(f"\nProcessant {nom_geo}...")
        df = construir_df(dades_const, dades_canc, geo_text)
        if df.empty:
            print(f"  Sense dades.")
            continue
        ws = wb.create_sheet(title=nom_geo)
        escriure_pestanya_dades(ws, df)
        print(f"  OK: {len(df)} periodes.")

    if len(wb.sheetnames) <= 1:
        print("ERROR: cap pestanya de dades creada.")
        sys.exit(1)

    wb.save(OUTPUT_FILE)
    print(f"Fitxer generat: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
